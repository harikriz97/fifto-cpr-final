"""
live/export_excel.py — Daily trade log export to Excel
=======================================================
Exports all trades for the day (paper_trades.csv + live_trades.csv)
to a formatted Excel file: data/reports/trades_YYYYMMDD.xlsx

Called automatically at EOD (15:20) by paper_trader.py.
Can also be run standalone: python live/export_excel.py
"""
from __future__ import annotations
import os
import sys
import logging
import pandas as pd
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
        GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E2530"
C_HEADER_FG   = "E2E8F0"
C_WIN_BG      = "0D2818"
C_WIN_FG      = "3FB950"
C_LOSS_BG     = "2D1111"
C_LOSS_FG     = "F85149"
C_ALT_BG      = "131720"
C_BASE_BG     = "0D0F14"
C_BASE_FG     = "C9D1D9"
C_BORDER      = "21262D"
C_ACCENT      = "58A6FF"
C_YELLOW      = "E3B341"
C_PURPLE      = "A78BFA"

AGENT_COLOURS = {
    "THOR":      (C_ACCENT,   "0D2137"),
    "HULK":      (C_WIN_FG,   "0D2818"),
    "IRON MAN":  (C_YELLOW,   "271D08"),
    "CAPTAIN":   (C_PURPLE,   "1A0D2E"),
    "CRT":       ("F97316",   "2D1500"),
    "MRC":       ("EC4899",   "2D0920"),
}


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)

def _font(hex_colour: str = C_BASE_FG, bold: bool = False,
          size: int = 9) -> Font:
    return Font(color=hex_colour, bold=bold, size=size, name="Calibri")

def _border() -> Border:
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=False)


def load_trades(target_date: str | None = None) -> pd.DataFrame:
    """
    Load and merge paper_trades.csv + live_trades.csv.
    target_date: 'YYYYMMDD' filter (defaults to today).
    """
    from config import DATA_DIR
    today = target_date or date.today().strftime("%Y%m%d")

    dfs = []
    for fname, ep_col, sym_col, src in [
        ("paper_trades.csv", "entry_price", "strategy", "paper_trader"),
        ("live_trades.csv",  "entry_price", "symbol",   "live_trader"),
    ]:
        path = os.path.join(DATA_DIR, fname)
        if not os.path.exists(path):
            continue
        try:
            df = pd.read_csv(path)
            df["_source"] = src
            # normalise date column
            df["date"] = df["date"].astype(str).str.replace("-", "").str[:8]
            df = df[df["date"] == today]
            if df.empty:
                continue
            # normalise column names
            df = df.rename(columns={ep_col: "entry_price", sym_col: "symbol_raw"})
            dfs.append(df)
        except Exception as e:
            logger.warning("load_trades %s: %s", fname, e)

    if not dfs:
        return pd.DataFrame()

    df = pd.concat(dfs, ignore_index=True, sort=False)

    # Ensure key columns exist
    for col, default in [
        ("signal",""), ("strategy",""), ("zone",""), ("opt",""),
        ("strike",0),  ("lots",1),  ("score",0),
        ("entry_price",0), ("exit_price",0), ("pnl",0),
        ("exit_reason",""), ("dte",0),
    ]:
        if col not in df.columns:
            df[col] = default

    df["pnl"]         = pd.to_numeric(df["pnl"],         errors="coerce").fillna(0)
    df["entry_price"] = pd.to_numeric(df["entry_price"], errors="coerce").fillna(0)
    df["exit_price"]  = pd.to_numeric(df["exit_price"],  errors="coerce").fillna(0)
    df["lots"]        = pd.to_numeric(df["lots"],         errors="coerce").fillna(1).astype(int)
    df["score"]       = pd.to_numeric(df["score"],        errors="coerce").fillna(0).astype(int)
    df["win"]         = (df["pnl"] > 0).astype(int)

    return df.reset_index(drop=True)


def export_daily_excel(target_date: str | None = None,
                        out_dir: str | None = None) -> str | None:
    """
    Export daily trade log to Excel.
    Returns output file path on success, None on failure.
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed — pip install openpyxl")
        return None

    from config import DATA_DIR, REPORT_DIR
    today     = target_date or date.today().strftime("%Y%m%d")
    out_dir   = out_dir or REPORT_DIR
    os.makedirs(out_dir, exist_ok=True)
    out_path  = os.path.join(out_dir, f"trades_{today}.xlsx")

    df = load_trades(today)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Sheet 1: Today's Trades ───────────────────────────────────────────────
    _write_trades_sheet(wb, df, today)

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    _write_summary_sheet(wb, df, today)

    # ── Sheet 3: All-time equity (cumulative) ─────────────────────────────────
    _write_equity_sheet(wb)

    wb.save(out_path)
    logger.info("Excel exported: %s", out_path)
    return out_path


def _write_trades_sheet(wb, df: pd.DataFrame, today: str):
    ws = wb.create_sheet("Trades")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_ACCENT

    # Header row
    headers = [
        "Time", "Agent", "Strategy", "Zone", "Opt", "Strike",
        "Entry", "Exit", "Exit Reason", "Lots", "Score", "DTE",
        "P&L (Rs.)", "Win",
    ]
    col_widths = [8, 10, 12, 14, 5, 8, 8, 8, 12, 6, 6, 5, 12, 5]

    header_fill = _fill(C_HEADER_BG)
    header_font = _font(C_HEADER_FG, bold=True, size=10)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = _align("center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 18

    # Title row above headers
    ws.insert_rows(1)
    title_text = f"FIFTO — Daily Trade Log  |  {today[:4]}-{today[4:6]}-{today[6:]}"
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(headers))
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.fill = _fill("0A0C10")
    tc.font = Font(color=C_ACCENT, bold=True, size=12, name="Calibri")
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 22

    if df.empty:
        ws.cell(row=3, column=1, value="No trades today.")
        ws.cell(row=3, column=1).font = _font(C_BASE_FG)
        return

    for ri, row in df.iterrows():
        r = ri + 3   # row 1=title, row 2=header, row 3+ = data
        pnl    = float(row.get("pnl", 0) or 0)
        is_win = pnl > 0
        sig    = str(row.get("signal", ""))
        agent_fg, agent_bg = AGENT_COLOURS.get(sig, (C_BASE_FG, C_ALT_BG))
        row_bg  = C_WIN_BG if is_win else (C_LOSS_BG if pnl < 0 else C_ALT_BG)
        pnl_fg  = C_WIN_FG if is_win else (C_LOSS_FG if pnl < 0 else C_BASE_FG)

        vals = [
            row.get("entry_time", row.get("exit_time", "")),
            sig,
            str(row.get("strategy", "")),
            str(row.get("zone", "")),
            str(row.get("opt", "")),
            int(row.get("strike", 0) or 0),
            float(row.get("entry_price", 0) or 0),
            float(row.get("exit_price",  0) or 0),
            str(row.get("exit_reason", "")),
            int(row.get("lots",  1) or 1),
            int(row.get("score", 0) or 0),
            int(row.get("dte",   0) or 0),
            pnl,
            "WIN" if is_win else ("LOSS" if pnl < 0 else "-"),
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _border()
            cell.alignment = _align("center" if ci in (1,5,9,14) else "right" if ci >= 6 else "left")
            cell.font = _font(C_BASE_FG)
            cell.fill = _fill(row_bg)

            # Special colours
            if ci == 2:   # Agent
                cell.fill = _fill(agent_bg)
                cell.font = _font(agent_fg, bold=True)
            if ci == 13:  # P&L
                cell.font = _font(pnl_fg, bold=True)
                cell.number_format = '#,##0.00'
            if ci == 14:  # Win/Loss
                cell.font = _font(C_WIN_FG if is_win else C_LOSS_FG, bold=True)
            if ci in (7, 8):   # Entry/Exit prices
                cell.number_format = '#,##0.00'

        ws.row_dimensions[r].height = 15

    # Freeze header rows
    ws.freeze_panes = "A3"


def _write_summary_sheet(wb, df: pd.DataFrame, today: str):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_WIN_FG
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16

    rows = []
    if not df.empty:
        total_pnl  = df["pnl"].sum()
        n_trades   = len(df)
        n_win      = int(df["win"].sum())
        n_loss     = n_trades - n_win
        wr         = round(n_win / n_trades * 100, 1) if n_trades else 0
        avg_win    = df[df["pnl"] > 0]["pnl"].mean() if n_win else 0
        avg_loss   = df[df["pnl"] < 0]["pnl"].mean() if n_loss else 0
        max_win    = df["pnl"].max() if n_trades else 0
        max_loss   = df["pnl"].min() if n_trades else 0
        by_agent   = df.groupby("signal")["pnl"].sum().to_dict()
        rows = [
            ("Date",            today[:4] + "-" + today[4:6] + "-" + today[6:]),
            ("Trades",          n_trades),
            ("Wins",            n_win),
            ("Losses",          n_loss),
            ("Win Rate",        f"{wr}%"),
            ("Total P&L",       f"Rs.{total_pnl:,.0f}"),
            ("Avg Win",         f"Rs.{avg_win:,.0f}"),
            ("Avg Loss",        f"Rs.{avg_loss:,.0f}"),
            ("Best Trade",      f"Rs.{max_win:,.0f}"),
            ("Worst Trade",     f"Rs.{max_loss:,.0f}"),
            ("",                ""),
            ("BY AGENT",        "P&L"),
        ] + [(f"  {k}", f"Rs.{v:,.0f}") for k, v in by_agent.items()]
    else:
        rows = [("Date", today), ("Trades", 0), ("Total P&L", "Rs.0")]

    # Title
    ws.merge_cells("A1:B1")
    tc = ws.cell(1, 1, "Summary")
    tc.fill = _fill("0A0C10")
    tc.font = Font(color=C_ACCENT, bold=True, size=12, name="Calibri")
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 22

    for ri, (label, val) in enumerate(rows, 2):
        lc = ws.cell(ri, 1, label)
        vc = ws.cell(ri, 2, val)
        is_sep  = label == ""
        is_hdr  = label in ("BY AGENT",)
        pnl_row = str(val).startswith("Rs.")
        lc.fill = vc.fill = _fill("0A0C10" if is_hdr else C_ALT_BG)
        lc.font = _font(C_ACCENT if is_hdr else C_BASE_FG, bold=is_hdr)
        pnl_col = C_WIN_FG if (pnl_row and "−" not in str(val) and "-" not in str(val)) \
                  else (C_LOSS_FG if pnl_row else C_BASE_FG)
        vc.font = _font(pnl_col, bold=is_hdr or label=="Total P&L")
        lc.border = vc.border = _border()
        lc.alignment = _align("left")
        vc.alignment = _align("right")
        ws.row_dimensions[ri].height = 15


def _write_equity_sheet(wb):
    """Cumulative equity from all historical trades."""
    ws = wb.create_sheet("Equity")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_YELLOW

    try:
        from config import DATA_DIR
        dfs = []
        for fname in ("paper_trades.csv", "live_trades.csv"):
            p = os.path.join(DATA_DIR, fname)
            if os.path.exists(p):
                d = pd.read_csv(p)
                d["pnl"] = pd.to_numeric(d["pnl"], errors="coerce").fillna(0)
                d["date"] = pd.to_datetime(d["date"].astype(str).str.replace("-","").str[:8],
                                           format="%Y%m%d", errors="coerce")
                dfs.append(d)
        if not dfs:
            ws.cell(1,1,"No historical data")
            return
        all_df = pd.concat(dfs).dropna(subset=["date"]).sort_values("date")
        daily  = all_df.groupby("date")["pnl"].sum().reset_index()
        daily["cum_pnl"] = daily["pnl"].cumsum()

        # Write headers
        ws.cell(1,1,"Date").fill = _fill(C_HEADER_BG)
        ws.cell(1,2,"Daily P&L").fill = _fill(C_HEADER_BG)
        ws.cell(1,3,"Cum P&L").fill = _fill(C_HEADER_BG)
        for ci in (1,2,3):
            ws.cell(1,ci).font = _font(C_HEADER_FG, bold=True)
            ws.cell(1,ci).alignment = _align("center")
            ws.cell(1,ci).border = _border()
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

        for ri, (_, row) in enumerate(daily.iterrows(), 2):
            ws.cell(ri,1, row["date"].strftime("%Y-%m-%d")).fill = _fill(C_ALT_BG)
            ws.cell(ri,2, round(row["pnl"],0)).fill = _fill(C_WIN_BG if row["pnl"]>=0 else C_LOSS_BG)
            ws.cell(ri,3, round(row["cum_pnl"],0)).fill = _fill(C_ALT_BG)
            ws.cell(ri,2).font = _font(C_WIN_FG if row["pnl"]>=0 else C_LOSS_FG)
            ws.cell(ri,3).font = _font(C_WIN_FG if row["cum_pnl"]>=0 else C_LOSS_FG, bold=True)
            for ci in (1,2,3):
                ws.cell(ri,ci).border = _border()
                ws.cell(ri,ci).alignment = _align("center")

        # Add line chart
        if len(daily) >= 2:
            chart = LineChart()
            chart.title  = "Cumulative P&L"
            chart.style  = 10
            chart.y_axis.title = "Rs."
            chart.x_axis.title = "Date"
            data_ref = Reference(ws, min_col=3, min_row=1, max_row=len(daily)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.series[0].graphicalProperties.line.solidFill = C_ACCENT
            chart.series[0].graphicalProperties.line.width = 20000
            ws.add_chart(chart, f"E2")

    except Exception as e:
        ws.cell(1,1,f"Equity chart error: {e}")


# ── CLI ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    target = sys.argv[1] if len(sys.argv) > 1 else None
    path   = export_daily_excel(target)
    if path:
        print(f"Exported: {path}")
        # Auto-open on Windows
        try:
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)
        except Exception:
            pass
    else:
        print("Export failed or no trades today.")
